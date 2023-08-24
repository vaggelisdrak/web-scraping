from tkinter import *
from tkinter import ttk
import tkinter
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import os
import pandas as pd
from twocaptcha import TwoCaptcha

root = Tk()
root.geometry('130x260')

def solveRecaptcha(sitekey, url, apikey):
    if apikey:
        api_key = os.getenv('APIKEY_2CAPTCHA', str(apikey))
        print('input given')
    else:
        api_key = os.getenv('APIKEY_2CAPTCHA', 'API_KEY')
        print('no input given')

    solver = TwoCaptcha(api_key)

    try:
        result = solver.recaptcha(
            sitekey=sitekey,
            url=url)

    except Exception as e:
        print(e)

    else:
        return result


def open_file(apikey,fromm,too,option):

    file = askopenfile(mode ='r+', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want. 
    #' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files
    print(file.name)
    wb = load_workbook(filename = file.name) # Load into openpyxl
    wb2 = wb.active

    #Whatever you want to do with the WorkSheet
    browser = webdriver.Chrome(ChromeDriverManager().install())
    #read excel data
    
    # read by default 1st sheet of an excel file
    dataframe1 = pd.read_excel(file.name)

    if option == 'ΑΦΜ':
        mylist1 = dataframe1['Α.Φ.Μ.'].tolist()
        
        print(mylist1)
        print('\n')

        mylist = []
        for i in mylist1:
            if len(str(i))<9:
                j = '0'+str(i)
                mylist.append(str(j))
            else:
                mylist.append(str(i))
        print('\n\n')
        print(mylist)
        search_by = "afm"

    elif option == 'ΕΠΩΝΥΜΙΑ':
        mylist1 = dataframe1['Επωνυμία Πελάτη'].tolist()
        mylist = []
        for i in mylist1:
            mylist.append(str(i))
        print('\n\n')
        print(mylist)
        search_by = "coName"
    else:
        mylist1 = dataframe1['Διακριτικός τίτλος'].tolist()
        mylist = []
        for i in mylist1:
            mylist.append(str(i))
        print('\n\n')
        print(mylist)
        search_by = "coName"

    #afms=0
    
    afms = int(fromm)-2
    if int(fromm)<=0 or afms<0:
        afms =0
    for i in mylist[afms:int(too)]:
        #search for AFM
        if afms<int(too):
            browser.get('https://www.businessregistry.gr/publicity/index')
        else:
            root.quit()
        print('excel line: ',afms+2)
        print('\nAFM: ',i)
        try:
            browser.find_element(By.ID,search_by).send_keys(str(i))
        except:
            root.quit()

        try:
            result = solveRecaptcha("6Ldqol4UAAAAAERYyWxbglpQTEO8NB_HLBVNAyuc","https://www.businessregistry.gr/publicity/index",apikey)
            code = result['code']
            print(code)
        except:
            continue

        WebDriverWait(browser, 1).until(
            EC.presence_of_element_located((By.ID, 'g-recaptcha-response'))
        )

        browser.execute_script(
            "document.getElementById('g-recaptcha-response').innerHTML = " + "'" + code + "'")

        #browser.find_element(By.CLASS_NAME, "submit btn btn-block btn-warning").click()
        try:
            browser.find_element(By.XPATH, "//*[@id='criteria_form']/ul/li[5]/button").click() #click sumbit after recaptcha
        except:
            browser.find_element(By.XPATH, "//*[@id='criteria_form']/ul/li[4]/button").click() #click sumbit after recaptcha

        if search_by == "afm":
            time.sleep(2)
        else:
            time.sleep(6)
        
        try:
            browser.find_element(By.CLASS_NAME,"tltp").click() #an den yparxoyn dedomeno proxora sto epomeno afm
        except:
            workbook = load_workbook(filename=file.name)
            sheet = workbook.active
            sheet['H'+str(int(afms)+2)] = 'Δεν βρεθηκαν αποτελεσματα'
            workbook.save(filename=file.name)
            afms+=1
            continue
        

        #browser.find_element(By.XPATH,"//td[2]/a").click()

        output_lst = []

        #gia ton proto pinaka-----------------------------------------------------------------
        for tr in browser.find_elements_by_xpath('//*[@id="home"]/div[1]/div/div/dl[1]'):
            try:
                tds = tr.find_elements_by_tag_name('dd')
            except:
                pass

        try:
            output_lst1 = [td.text for td in tds]
            #print('output_lst',output_lst1)
        except:
            try:
                for tr in browser.find_elements_by_xpath("/html/body/div[1]/div[2]/div/div/div/div/div/div/div/div[3]/div/div/div/table/tbody"):
                    try:
                        tds = tr.find_elements_by_tag_name('td') 
                    except:
                        pass
                
                output = [td.text for td in tds]
                print(output)
                workbook = load_workbook(filename=file.name)
                sheet = workbook.active
                k=0
                for i in output:
                    column= ['H','I','J','L','M','O']
                    try:
                        sheet[column[k]+str(int(afms)+2)] = str(i)
                    except:
                        pass
                    k+=1
                workbook.save(filename=file.name)
                afms+=1
                continue
            except:
                workbook = load_workbook(filename=file.name)
                sheet = workbook.active
                sheet['H'+str(int(afms)+2)] = 'Παρουσιαστηκε σφαλμα1'
                workbook.save(filename=file.name)
                afms+=1
                continue

        #gia ton deytero pinaka----------------------------------------------------------
        for tr in browser.find_elements_by_xpath('//*[@id="home"]/div[1]/div/div/dl[2]'):
            try:
                tds = tr.find_elements_by_tag_name('dd')
            except:
                pass
        
        try:
            output_lst2 = [td.text for td in tds]
            output_lst2.pop(0)
            print('output_lst2',output_lst2)
        except:
            workbook = load_workbook(filename=file.name)
            sheet = workbook.active
            sheet['H'+str(int(afms)+2)] = 'Παρουσιαστηκε σφαλμα2'
            workbook.save(filename=file.name)
            afms+=1
            continue


        #gia ton trito pinaka / dioikhsh----------------------------------------------------------
        output_lst3 = []
        try:
            ben = browser.find_elements_by_xpath('//*[@id="home"]/div[1]/div/div/dl[4]')
            print(ben)
            for tr in browser.find_elements_by_xpath('//*[@id="home"]/div[1]/div/div/dl[4]'):
                tdss = tr.find_elements_by_tag_name('dd')
                print('tdss1',tdss)
                output_lst3 = [td.text for td in tdss]
        except:
            try:
                for tr in browser.find_elements_by_xpath('//*[@id="home"]/div[1]/div/div/dl[5]'):
                    tdss = tr.find_elements_by_tag_name('dd')
                    print('tdss2',tdss)
                    output_lst3 = [td.text for td in tdss]
            except:
                output_lst3 = []

        #output_lst3 = [td.text for td in tdss]
        #output_lst3.pop(0)
        print('output_lst3',output_lst3)
        '''try:
            output_lst3 = [td.text for td in tdss]
            #output_lst3.pop(0)
            print('output_lst3',output_lst3)
        except:
            workbook = load_workbook(filename=file.name)
            sheet = workbook.active
            sheet['H'+str(int(afms)+2)] = 'Παρουσιαστηκε σφαλμα3'
            workbook.save(filename=file.name)
            afms+=1
            continue
        '''
        
        lst = output_lst1 + output_lst2 + output_lst3

        print(lst)
        # Start by opening the spreadsheet and selecting the main sheet
        workbook = load_workbook(filename=file.name)
        sheet = workbook.active

        # Write what you want into a specific cell
        k=0
        for i in lst:
            column= ['H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y']
            try:
                sheet[column[k]+str(int(afms)+2)] = str(i)
            except:
                pass
            k+=1

        # Save the spreadsheet
        workbook.save(filename=file.name)
        afms+=1

#GUI

user_input = tkinter.StringVar(root)
fromm = tkinter.StringVar(root)
too = tkinter.StringVar(root)

l1 = Label(root, text="Enter API KEY")
e1 = Entry(root, bd =5,textvariable=user_input)
#w2 = Scale(root, from_ = 1, to = 1000, orient = HORIZONTAL)
#w3 = Scale(root, from_ = 1, to = 1000, orient = HORIZONTAL)
l2 = Label(root, text="From")
e2 = Entry(root, bd =5,textvariable=fromm)

l3 = Label(root, text="To")
e3 = Entry(root, bd =5,textvariable=too)

options = ['ΑΦΜ','ΕΠΩΝΥΜΙΑ','ΔΙΑΚΡΙΤΙΚΟΣ']
l4 = ttk.Combobox(root, value=options,width=15)
l4.grid(row=3,column=0)
l4.current(0)

btn = Button(root, text ='Open excel file', command = lambda: open_file(user_input.get(),fromm.get(),too.get(), l4.get()))

l1.grid(row = 0, column = 0,  pady = 2)
e1.grid(row = 1, column = 0,  pady = 2)

l2.grid(row = 2, column = 0,  pady = 2 )
e2.grid(row = 3, column = 0,  pady = 2)

l3.grid(row = 4, column = 0,  pady = 2)
e3.grid(row = 5, column = 0,  pady = 2)

l4.grid(row=6,column=0,pady=10)
btn.grid(row=7, column = 0,  pady = 20)

try:
    mainloop()
except:
    pass
