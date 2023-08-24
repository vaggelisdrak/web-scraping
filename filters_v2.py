from tkinter import *
import tkinter
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
from seleniumbase import Driver

browser = Driver(uc=True)
#root.geometry('275x85')
file = askopenfile(mode ='r+', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want. 
print(file.name)

def open_website():
    browser.get('https://publicity.businessportal.gr/')

# Function to check for the error message and reload the page
def check_and_reload(driver):
    error_message = "net::ERR_INCOMPLETE_CHUNKED_ENCODING"
    console_logs = driver.get_log("browser")
    
    for log in console_logs:
        if error_message in log["message"]:
            print("Detected ERR_INCOMPLETE_CHUNKED_ENCODING error. Reloading...")
            driver.refresh()
            return True
    
    return False

def scrape_data():
    #Start scraping the data
    time.sleep(1)

    #click search
    try:
        browser.find_element(By.XPATH,'//*[@class="MuiBox-root css-1ofqig9"]/div[5]/div/div[3]/button').click()
    except:
        pass

    time.sleep(2)

    # Find all the div elements with the specified class
    div_elements = browser.find_elements(By.XPATH,'//*["MuiCardContent-root.css-1qw96cp"]')

    page = 1
    lst_of_dict = []

    while True:
        time.sleep(2)
        #button_class = '//*[@id="vertical-tabpanel-0"]/div/div/div/div[1]/nav/ul/li[9]/button'
       # Find and click the last li button using XPath
        
        print("\npage number: ",page)
        page+=1
        for i in range(1,11):
            if i==1:
                try:
                    element = browser.find_element(By.XPATH,'//*[@class="MuiCardContent-root css-1qw96cp"]/a')
                    element.click()
                    title = element.text
                except:
                    try:
                        check_and_reload(browser)
                        wait = WebDriverWait(browser, 2)  # max 10 seconds timeout
                        element = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@class="MuiPaper-root MuiPaper-elevation MuiPaper-rounded MuiPaper-elevation1 MuiCard-root css-s18byi"][' + str(i) + ']/div/a/p')))
                        element.click()
                        title = element.text
                    except:      
                        # Refresh the page
                        check_and_reload(browser)
                        wait = WebDriverWait(browser, 2)  # max 10 seconds timeout
                        element = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@class="MuiPaper-root MuiPaper-elevation MuiPaper-rounded MuiPaper-elevation1 MuiCard-root css-s18byi"][' + str(i) + ']/div/a/p')))
                        element.click()
                        title = element.text
                scrape_data_details(lst_of_dict,title)
            else:
                try:
                    check_and_reload(browser)
                    wait = WebDriverWait(browser, 2)  # max 10 seconds timeout
                    element = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@class="MuiPaper-root MuiPaper-elevation MuiPaper-rounded MuiPaper-elevation1 MuiCard-root css-s18byi"][' + str(i) + ']/div/a/p')))
                    element.click()
                    title = element.text
                except:
                    # Refresh the page
                    check_and_reload(browser)
                    wait = WebDriverWait(browser, 2)  # max 10 seconds timeout
                    element = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@class="MuiPaper-root MuiPaper-elevation MuiPaper-rounded MuiPaper-elevation1 MuiCard-root css-s18byi"][' + str(i) + ']/div/a/p')))
                    element.click()
                    title = element.text
                scrape_data_details(lst_of_dict,title)

        try:  
            time.sleep(1)
            button_xpath = '//*[@id="vertical-tabpanel-0"]/div/div/div/div[1]/nav/ul/li[last()]/button'
            button = browser.find_element(By.XPATH, button_xpath)
            button.click()
            time.sleep(1)

        except:
            #an den yparxoyn alles selides stamata
            print('\nno more pages')
            break

    print(lst_of_dict)
    
def scrape_data_details(lst_of_dict,title):
    time.sleep(2)

    #get data from table
    get_url = browser.current_url
    print("The current url is: "+str(get_url))

    # Navigate to the URL
    #browser.get(get_url)

    # Get the page source using Selenium
    page_source = browser.page_source

    # Parse the page source with BeautifulSoup
    soup = BeautifulSoup(page_source, 'html.parser')

    # Find all rows within the table body
    table_rows = soup.select('tbody.MuiTableBody-root tr.MuiTableRow-root')

    # Initialize a dictionary to store the extracted div text
    div_texts_dict = {}
    lst = []

    # get title
    time.sleep(2)
    if not title:
        try:
            h4_element = soup.find('h4', class_='MuiTypography-root MuiTypography-h4 css-1xvinid')
            text = h4_element.get_text()
        except:
            try:
                element = browser.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[1]/div/h4')
                text = element.text
            except:
                pass

        print(text)
        div_texts_dict["Eπωνυμία"] = text
        lst.append(text)
    else:
        text = title
        print(text)
        div_texts_dict["Eπωνυμία"] = text
        lst.append(text)


    # Loop through each row and extract the text from columns 1 and 2
    i=0
    for row in table_rows:
        i+=1
        if i==1: continue #skip first row
        

        columns = row.find_all('td', class_='MuiTableCell-root')
        if len(columns) >= 2:
            column1_text = columns[0].get_text()
            column2_text = columns[1].get_text()
            if i<=12: #first table with basic info
                div_texts_dict[column1_text] = column2_text
                lst.append(column2_text)

            else: 
                if column1_text == "E-mail" or column1_text == "Τηλέφωνο":
                    div_texts_dict[column1_text] = column2_text
                    lst.append(column2_text)
                elif len(column1_text) >= 3:
                    if column1_text[2] == '.':
                        div_texts_dict[column1_text] = column2_text
                        lst.append(column2_text)
                else:
                    pass
                
    # Print the dictionary of div texts
    print(div_texts_dict)
    lst_of_dict.append(div_texts_dict)

    # Start by opening the spreadsheet and selecting the main sheet
    workbook = load_workbook(filename=file.name)
    sheet = workbook.active

    # Find the last row in the Excel sheet
    last_row = sheet.max_row + 1

    # Write what you want into a specific cell in the excel file
    # Dictionary keys in the same order as the columns in the sheet
    dictionary_keys = list(div_texts_dict.keys())  # Add all keys here

    # Write dictionary values to corresponding cells in the last row
    for column_index, key in enumerate(dictionary_keys, start=1):
        cell = sheet.cell(row=last_row, column=column_index)
        cell.value = div_texts_dict.get(key, '')
    

    # Save the spreadsheet
    workbook.save(filename=file.name)
    

    browser.get('https://publicity.businessportal.gr/')

        
#GUI

root = Tk()

user_input = tkinter.StringVar(root)
fromm = tkinter.StringVar(root)
too = tkinter.StringVar(root)

l1 = Label(root, text="1-Click to Open website")
l2 = Label(root, text="2-Enter filters and when you\n are ready click continue")

btn1 = Button(root, text ='Open', command = lambda: open_website())
btn2 = Button(root, text ='Continue', command = lambda: scrape_data())

l1.grid(row = 0, column = 0,  pady = 5)
btn1.grid(row = 0, column = 1,  pady = 5)

l2.grid(row = 1, column = 0,  pady = 5)
btn2.grid(row=1, column = 1,  pady = 20)

try:
    mainloop()
except:
    pass