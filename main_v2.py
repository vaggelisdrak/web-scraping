from tkinter import *
from tkinter import ttk
import tkinter
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
import os
import pandas as pd
from twocaptcha import TwoCaptcha
from seleniumbase import Driver

root = Tk()
root.geometry('130x260')

# Function to check for the error message and reload the page
def check_and_reload(driver):
    error_message = "net::ERR_INCOMPLETE_CHUNKED_ENCODING"
    console_logs = driver.get_log("browser")
    
    for log in console_logs:
        if error_message in log["message"]:
            print("Detected ERR_INCOMPLETE_CHUNKED_ENCODING error. Reloading...")
            driver.refresh()
            return True
    
    return False

def choose_option(option,dataframe1,mylist):
    if option == 'ΑΦΜ':
        mylist1 = dataframe1['Α.Φ.Μ.'].tolist()
        
        print(mylist1)
        print('\n')

        for i in mylist1:
            if len(str(i))<9:
                j = '0'+str(i)
                mylist.append(str(j))
            else:
                mylist.append(str(i))
        print('\n\n')
        print(mylist)

    elif option == 'ΕΠΩΝΥΜΙΑ':
        mylist1 = dataframe1['Επωνυμία Πελάτη'].tolist()
        for i in mylist1:
            mylist.append(str(i))
        print('\n\n')
        print(mylist)
    else:
        mylist1 = dataframe1['Διακριτικός τίτλος'].tolist()
        for i in mylist1:
            mylist.append(str(i))
        print('\n\n')
        print(mylist)
    
def save_data(file,afms,div_texts_dict,lst):
    # Start by opening the spreadsheet and selecting the main sheet
    workbook = load_workbook(filename=file.name)
    sheet = workbook.active
    # Write what you want into a specific cell in the excel file
    k=0
    for i in lst:
        column= ['H','I','J','K','L','M','N','O','P','Q','R','S']
        
        value = list(div_texts_dict.keys())[list(div_texts_dict.values()).index(i)]

        if value == "E-mail":
            sheet['T'+str(int(afms)+2)] = str(i)
        elif value == "Τηλέφωνο":
            sheet['U'+str(int(afms)+2)] = str(i)
        elif value.startswith("56."):
            sheet['V'+str(int(afms)+2)] = str(i)
        elif value == 'Επωνυμία':
            sheet['W'+str(int(afms)+2)] = str(i)
        else:
            try:
                sheet[column[k]+str(int(afms)+2)] = str(i)
            except:
                continue
        k+=1

    # Save the spreadsheet
    workbook.save(filename=file.name)
    pass

def open_file(apikey,fromm,too,option):

    file = askopenfile(mode ='r+', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want. 
    print(file.name)
    wb = load_workbook(filename = file.name) # Load into openpyxl
    wb2 = wb.active

    #Whatever you want to do with the WorkSheet

    #browser = webdriver.Chrome(ChromeDriverManager(version="116.0.5845.96").install())
    browser = Driver(uc=True)

    #read excel data
    
    # read by default 1st sheet of an excel file
    dataframe1 = pd.read_excel(file.name)
    mylist = []

    # choose scraping option (by afm, by eponymia, by diakritikos titlos)
    choose_option(option,dataframe1,mylist)

    #Start scraping the data

    #afms=0
    
    afms = int(fromm)-2
    if int(fromm)<=0 or afms<0:
        afms =0

    if afms<int(too):
        browser.get('https://publicity.businessportal.gr/')
        time.sleep(2)
        check_and_reload(browser)
    else:
        root.quit()

    for i in mylist[afms:int(too)]:
        #search for AFM
        
        print('excel line: ',afms+2)
        print('AFM: ',i)

        time.sleep(1)
        browser.find_element(By.XPATH,"//input").clear()
        browser.find_element(By.ID,"AutocompleteSearchItem").clear()

        try:
            browser.find_element(By.ID,"AutocompleteSearchItem").click()
            browser.find_element(By.XPATH,'//*[@class="MuiAutocomplete-endAdornment css-2iz2x6"]/button').click() 
        except:
            pass
        
        #grapse sto search input
        browser.find_element(By.ID,"AutocompleteSearchItem").send_keys(str(i))

        #click search button
        browser.find_element(By.XPATH,'//*[@class="MuiButtonBase-root MuiIconButton-root MuiIconButton-colorPrimary MuiIconButton-sizeMedium css-1harbtz"]').click() 
        time.sleep(1)

        
        #phgaine sthn selida me ta details
        try:
            #browser.get('https://publicity.businessportal.gr/')
            time.sleep(2)
            check_and_reload(browser)
            wait = WebDriverWait(browser, 3)  # max 3 seconds timeout
            element = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@class="MuiCardContent-root css-1qw96cp"]/a/p')))
            element.click()
            title = element.text
        except:
            check_and_reload(browser)
            #an den yparxoyn dedomeno proxora sto epomeno afm
            workbook = load_workbook(filename=file.name)
            sheet = workbook.active
            sheet['H'+str(int(afms)+2)] = 'Δεν βρεθηκαν αποτελεσματα'
            print("Δεν βρεθηκαν αποτελεσματα")
            workbook.save(filename=file.name)
            afms+=1
            continue

        time.sleep(2)

        #get data from table
        get_url = browser.current_url
        print("The current url is: "+str(get_url))

        # Navigate to the URL
        browser.get(get_url)

        # Get the page source using Selenium
        page_source = browser.page_source

        # Parse the page source with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find all rows within the table body
        table_rows = soup.select('tbody.MuiTableBody-root tr.MuiTableRow-root')

        # Initialize a dictionary to store the extracted div text
        div_texts_dict = {}
        lst = []

        div_texts_dict['Επωνυμία'] = title
        lst.append(title)
        # Loop through each row and extract the text from columns 1 and 2
        i=0
        for row in table_rows:
            i+=1
            if i==1: continue #skip first row
            

            columns = row.find_all('td', class_='MuiTableCell-root')
            if len(columns) >= 2:
                column1_text = columns[0].get_text()
                column2_text = columns[1].get_text()
                if i<=12: #first table with basic info
                    div_texts_dict[column1_text] = column2_text
                    lst.append(column2_text)

                else: 
                    if column1_text == "E-mail" or column1_text == "Τηλέφωνο":
                        div_texts_dict[column1_text] = column2_text
                        lst.append(column2_text)
                    elif column1_text.startswith("56."):
                        div_texts_dict[column1_text] = column2_text
                        lst.append(column2_text)
                    else:
                        pass
                    
        # Print the dictionary of div texts
        print(div_texts_dict)

        #save the data to excel file
        save_data(file,afms,div_texts_dict,lst)
        
    
        #go to previous page
        browser.get('https://publicity.businessportal.gr/')
        time.sleep(2)
        afms+=1

#GUI

user_input = tkinter.StringVar(root)
fromm = tkinter.StringVar(root)
too = tkinter.StringVar(root)

l2 = Label(root, text="From")
e2 = Entry(root, bd =5,textvariable=fromm)

l3 = Label(root, text="To")
e3 = Entry(root, bd =5,textvariable=too)

options = ['ΑΦΜ','ΕΠΩΝΥΜΙΑ','ΔΙΑΚΡΙΤΙΚΟΣ']
l4 = ttk.Combobox(root, value=options,width=15)
l4.grid(row=3,column=0)
l4.current(0)

btn = Button(root, text ='Open excel file', command = lambda: open_file(user_input.get(),fromm.get(),too.get(), l4.get()))

l2.grid(row = 2, column = 0,  pady = 2 )
e2.grid(row = 3, column = 0,  pady = 2)

l3.grid(row = 4, column = 0,  pady = 2)
e3.grid(row = 5, column = 0,  pady = 2)

l4.grid(row=6,column=0,pady=10)
btn.grid(row=7, column = 0,  pady = 20)

try:
    mainloop()
except:
    pass